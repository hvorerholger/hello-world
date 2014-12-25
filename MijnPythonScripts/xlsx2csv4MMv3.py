# xlsx to csv.py
#v2 change to argparser
#v3 implement functions invoked under if __name__ == '__main__'
#v4 implement try: except: logic to handle data input errors/ valid inputfilename/ already processed file
# implement from future import so that script could evt run in v2.4
# implement virtualenv/ explore
# implement Realized internal contracts ING LUX per Month dd, yyyy
# combine into one script
# make script location agnostic with mapping file style almSwapUpload


import os,csv,calendar,re,decimal,shutil
from openpyxl import load_workbook, Workbook, cell, styles
import argparse


def parseCmdLineInput():
    import argparse
    parser = argparse.ArgumentParser(description='convert xlsx File to csv File')
    parser.add_argument("-I", dest="input_file", required=True)
    args = parser.parse_args()
    return args.input_file

def validateFileName(myFileName):
    import re
    try:  #assert is not meant to validate user input!!!
        assert (re.search\
                (r'(?P<prefix>^replication)(?P<delimiter>[\s|_|-])(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]csv$)',\
                 myFileName,\
                 re.IGNORECASE))\
                 is not None       
    except AssertionError:
        print ('Srry invalid filename: ', '\t', myFileName, '\t\t', 'expecting: replication_YYYYMM.csv') # gaat dit naar stdout?
        return False
    else:
        searchObj = re.search(r'(?P<prefix>^replication)(?P<delimiter>[\s|_|-])(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]csv$)', myFileName, re.IGNORECASE)
        print ('BINGO! valid filename: ', '\t', myFileName, '\t','year: ',searchObj.group('year'),'   month: ',searchObj.group('month'))
        return True

def processXls2Csv(myxlsFile):
    import os,csv,calendar,re,decimal,shutil
    from openpyxl import load_workbook, Workbook, cell, styles

    inputFileName_original = re.findall('Replication.+\.xlsx', myxlsFile)[0]   #?? unsave -- should assert it is not null!! empty list
    inputFileName_new = "PROGNOSIS_" + inputFileName_original
    inputFile_new = myxlsFile.replace(inputFileName_original,inputFileName_new) #?? why is this// kan dit niet eenvoudiger??
    shutil.copyfile(myxlsFile, inputFile_new)
            
    wb = load_workbook(inputFile_new)
    try:
            ws_start = wb['Sheet1']
    except KeyError as e:
            print(e)
            print(wb.get_sheet_names())
            exit()

    for sheetName in wb.get_sheet_names():
            if sheetName == 'trade':
                    ws_trade = wb.get_sheet_by_name(sheetName)
                    wb.remove_sheet(ws_trade)
                    break
            
    ws_trade = wb.create_sheet()
    ws_trade.title = 'trade'

    csvFile = myxlsFile.replace(".xlsx",".csv").replace(".xls","csv")
    csvOpen = open(csvFile,'w',newline="")
    csvData = csv.writer(csvOpen)

    #---------------extract date from file name--------------
    yearMonth = re.findall('Replication.+\.xlsx', myxlsFile)[0][12:18]
    year = yearMonth[0:4]
    month = yearMonth[4:6]
    day = calendar.monthrange(int(year),int(month))[1]
    date = yearMonth + "%d" %(day)

    #-------------------write the headers-------------#
    headers = ["TradeDate","MMType","Ccy","Start","End","Notional","Index","Rate","Cpty","Company","Desk","Book"]
    ws_trade.append(headers)
    csvData.writerow(headers)
    ws_start['B17'].value = 'TOT'
    ws_start['A19'].value = 'InvCumGap TOT'
    ws_start['A20'].value = 'InvCumGap PR'

    #-------------------set number format----------------------------#
    _numberFormat = ws_start['D3'].number_format

    maxCol = ws_start.get_highest_column()
    for icol in range(5,maxCol+1): #range(5,89) is from column E to column CJ
            #-------------------generate total notional---------------#
            cell = ws_start.cell(row=17,column=icol)
            cell.value = "=SUM(" + ws_start.cell(row=3,column=icol).coordinate + "," + ws_start.cell(row=6,column=icol).coordinate + "," + ws_start.cell(row=9,column=icol).coordinate + ")"
            cell.number_format = _numberFormat
            #-------------------generate trade scenarios---------------#
            for irow in range(3,10,3):
                    cell = ws_start.cell(row=irow,column=icol)
                    if cell.value == 0:
                            continue
                    elif cell.value < 0:
                            MMType = "DEPOSIT"
                    else:
                            MMType = "LOAN"
                    endDate = ws_start.cell(row=2,column=icol).value
                    notional = "{0:.2f}".format(abs(ws_start.cell(row=irow,column=icol).value))
                    rate = "{0:.4f}".format(ws_start.cell(row=irow+1,column=icol).value/100)
                    book = ws_start.cell(row=irow,column=2).value
                    if book == 'PR':
                            lubook = 'LUMMPB'
                    elif book == 'RS':
                            lubook = 'LUMMRE'
                    elif book == 'WS':
                            lubook = 'LUMMCB'
                    else:
                            continue
                    row = [date,MMType,"EUR",date,endDate,notional,"FIXED",rate,"INTEINT","INGLU","LUDESK",lubook]
                    ws_trade.append(row)
                    csvData.writerow(row)
    del csvData
    csvOpen.close()

    #-------------------generate InvCumGap profile---------------#
    for icol in range(5,maxCol+1): #range(5,89) is from column E to column CJ
            cell = ws_start.cell(row=19,column=icol)
            cell.value = "=SUM(" + ws_start.cell(row=17,column=icol).coordinate + ":" + ws_start.cell(row=17,column=maxCol).coordinate  + ")"
            cell.number_format = _numberFormat
            cell = ws_start.cell(row=20,column=icol)
            cell.value = "=SUM(" + ws_start.cell(row=3,column=icol).coordinate + ":" + ws_start.cell(row=3,column=maxCol).coordinate  + ")"
            cell.number_format = _numberFormat

    wb.save(inputFile_new)

def main():
    parseCmdLineInput()
    validate_inputFile() #valid name, file exists, >0Kb, readable, content-checks, duplicate processing check (outfile already exists)
    processXls2Csv(xlsFile)
    archiveProcessedFile()
    #--
    emailPrognosis()
    extractRealized()
    emailRealized()


if __name__ == '__main__':
    #assert validateFileName('replication_201411.csv')
    #assert validateFileName('zomaarEenFileName.txt')
    #import os.path
    #assert os.path.isfile('PROGNOSIS_replication-201401.xlsx')
    xlsFile=parseCmdLineInput()
    processXls2Csv(xlsFile)
    assert os.path.isfile('PROGNOSIS_Replication_201411.xlsx')
    assert os.path.isfile('Replication_201411.csv')       
