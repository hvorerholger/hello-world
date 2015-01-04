import os
import datetime
from openpyxl import cell, load_workbook, styles, Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.drawing import Image
from openpyxl.comments import Comment

def spielerei():
    print(os.getcwd())
    wb = load_workbook('Replication-201401.xlsx')
    print(wb.get_sheet_names())
    ws2 = wb.get_sheet_by_name('Sheet2')
    ws2['A2'] = 999
    print ("{0:.2f}".format(ws2['A2'].value))

def writeWorkbook():
    wb = Workbook()
    dest_filename = r'pyxl_outfile.xlsx'
    ws1 = wb.create_sheet()
    ws1.title = 'pi'
    ws1['A1'] = 3.54
    ws2 = wb.create_sheet()
    ws2.title = 'ranges'
    for col_idx in range(1,10):
        col = get_column_letter(col_idx)
        for row in range(1,10):
            ws2.cell('%s%s'%(col,row)).value = '%s%s' % (col,row)
    wb.save(filename = dest_filename)

def readExistingWorkbook():    
    wb = load_workbook(filename = 'pyxl_outfile.xlsx')
    sheet_ranges = wb['ranges']
    print(sheet_ranges['D9'].value)
    sheet_pi = wb['pi']
    print(sheet_pi['A1'].value)

def using_number_formats():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = datetime.datetime(2010, 7, 21)
    print (ws['A1'].style.number_format.format_code) #returns 'yyyy-mm-dd'
    #set percentage using a string followed by the percent sign
    ws['B1'] = '3.14%'
    print (ws['B1'].value) #return 0.0314
    print (ws ['B1'].style.number_format_code) # returns '0%'

def using_formulae():
    wb = Workbook()
    ws = wb.active
    #add a simple formula
    ws['A1'] = "=SUM(1,1)"
    wb.save("formula.xlsx")

def merge_unmerge_cells():
    ws.merge_cells('A1:B1')
    ws.unmerge_cells('A1:B1')
    # or
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
    ws.unmerge_cells(start_row=2,start_column=1,end_row=2,end_column=4)

def inserting_an_image():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'You should see three logos below'
    ws['A2'] = 'Resize the rows and cells to see anchor differences'

    # create image instances
    img = Image('logo.png')
    img2 = Image('logo.png')
    img3 = Image('logo.png')

    # place image relative to top left corner of spreadsheet
    img.drawing.top = 100
    img.drawing.left = 150

    # the top left offset needed to put the image
    # at a specific cell can be automatically calculated
    img2.anchor(ws['D12'])
    (('D', 12), ('D', 21))

    # one can also position the image relative to the specified cell
    # this can be advantageous if the spreadsheet is later resized
    # (this might not work as expected in LibreOffice)
    img3.anchor(ws['G20'], anchortype='oneCell')
    ((6, 19), None)

    # afterwards one can still add additional offsets from the cell
    img3.drawing.left = 5
    img3.drawing.top = 5

    # add to worksheet
    ws.add_image(img)
    ws.add_image(img2)
    ws.add_image(img3)
    wb.save('logo.xlsx')

def fold_columns_outline():
    wb = openpyxl.Workbook(True)
    ws = wb.create_sheet()
    ws.column_dimensions.group('A','D', hidden=True)
    wb.save('group.xlsx')

def create_chart():
    wb = Workbook()
    ws = wb.active
    for i in range(10):
        ws.append([i])

    from openpyxl.charts import BarChart, Reference, Series
    values = Reference(ws, (1, 1), (10, 1))
    series = Series(values, title="First series of values")
    chart = BarChart()
    chart.append(series)
    ws.add_chart(chart)
    wb.save("SampleChart.xlsx")

def add_comment_to_cell():
    wb = Workbook()
    ws = wb.active
    comment = ws["A1"].comment
    comment = Comment('This is the comment text', 'Comment Author')
    comment.text
    'This is the comment text'
    comment.author
    'Comment Author'

def styles():
    from styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
    s = Style(font=Font(name='Calibri',
                       size=11,
                       bold=False,
                       italic=False,
                       vertAlign=None,
                       underline='none',
                       strike=False,
                       color='FF000000'),
             fill=PatternFill(fill_type=None,
                       start_color='FFFFFFFF',
                       end_color='FF000000'),
             border=Border(left=Side(border_style=None,
                                      color='FF000000'),
                            right=Side(border_style=None,
                                       color='FF000000'),
                            top=Side(border_style=None,
                                     color='FF000000'),
                            bottom=Side(border_style=None,
                                        color='FF000000'),
                            diagonal=Side(border_style=None,
                                          color='FF000000'),
                            diagonal_direction=0,
                            outline=Side(border_style=None,
                                         color='FF000000'),
                            vertical=Side(border_style=None,
                                          color='FF000000'),
                            horizontal=Side(border_style=None,
                                           color='FF000000')),
           alignment=Alignment(horizontal='general',
                               vertical='bottom',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=False,
                               indent=0),
           number_format='General',
           protection=Protection(locked='inherit',
                                 hidden='inherit'))

def redefine_style():
    from openpyxl.styles import Font, Color
    from openpyxl.styles import colors

    ft = Font(color=colors.RED)
    s1 = Style(font=ft, number_format='0%')
    s2 = Style(font=ft, number_format='dd-mm-yyyy')

def copy_style():
    from openpyxl.styles import Font, Style
    arial = Font(name='Arial', size=14)
    tahoma = Font(name='Tahoma')
    s1 = Style(font=arial, number_format='0%')
    s2 = s1.copy(font=tahoma)
    s2.font.name
    s2.number_format

def apply_style():
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Style
    #apply to cell
    wb = Workbook()
    ws = wb.active
    c = ws['A1']
    c.style = Style()
    #apply to colums and rows
    col = ws.column_dimensions['A']
    col.style = Style()
    row = ws.row_dimensions[1]
    row.style = Style()

def conditional_formatting():
    from openpyxl import Workbook
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.formatting import ColorScaleRule, CellIsRule, FormulaRule

    wb = Workbook()
    ws = wb.active

    # Create fill
    redFill = PatternFill(start_color='FFEE1111',
                   end_color='FFEE1111',
                   fill_type='solid')
    # Add a two-color scale
    # add2ColorScale(range_string, start_type, start_value, start_color, end_type, end_value, end_color)
    # Takes colors in excel 'FFRRGGBB' style.
    ws.conditional_formatting.add('A1:A10',
                ColorScaleRule(start_type='min', start_color=Color('FFAA0000'),
                              end_type='max', end_color=Color('FF00AA00'))
                              )

    # Add a three-color scale
    ws.conditional_formatting.add('B1:B10',
                   ColorScaleRule(start_type='percentile', start_value=10, start_color=Color('FFAA0000'),
                               mid_type='percentile', mid_value=50, mid_color=Color('FF0000AA'),
                               end_type='percentile', end_value=90, end_color=Color('FF00AA00'))
                                 )

    # Add a conditional formatting based on a cell comparison
    # addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
    # Format if cell is less than 'formula'
    ws.conditional_formatting.add('C2:C10',
                CellIsRule(operator='lessThan', formula=['C$1'], stopIfTrue=True, fill=redFill))

    # Format if cell is between 'formula'
    ws.conditional_formatting.add('D2:D10',
                CellIsRule(operator='between', formula=['1','5'], stopIfTrue=True, fill=redFill))

    # Format using a formula
    ws.conditional_formatting.add('E1:E10',
                FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))

    # Aside from the 2-color and 3-color scales, format rules take fonts, borders and fills for styling:
    myFont = Font()
    myBorder = Border()
    ws.conditional_formatting.add('E1:E10',
                FormulaRule(formula=['E1=0'], font=myFont, border=myBorder, fill=redFill))

    # Custom formatting
    # There are many types of conditional formatting - it's possible to add additional types directly:
    ws.conditional_formatting.add('E1:E10',
                {'type': 'expression', 'dxf': {'fill': redFill},
                 'formula': ['ISBLANK(E1)'], 'stopIfTrue': '1'})

    # Before writing, call setDxfStyles before saving when adding a conditional format that has a font/border/fill
    ws.conditional_formatting.setDxfStyles(wb)
    wb.save("testcondformat.xlsx")


    
if __name__ == "__main__":
    #spielerei()
    #writeWorkbook()
    #readExistingWorkbook()
    #create_chart()
    conditional_formatting()
    
