# xlsx to csv.py
# v2 change to argparser
# v3 implement functions invoked under if __name__ == '__main__'
# v4 implement try: except: logic
# v5 replace row/col references by cell coordinates
# check final xml outputs with 2 consecutive scripts run
# return args global for parseCmdLineInput \
# with defaults for optional arguments -D importdir and -A archivedir
# combine 2 scripts into one script
# implicit: output dir = script directory where also find xlsx input File
# implement from future import so that script could evt run in v2.4
# implement virtualenv/ explore
# implement Realized internal contracts ING LUX per Month dd, yyyy
# make script location agnostic with mapping file style almSwapUpload
# nose, pydev, coverage
# open questions:
# does Support run almSwapUpload each month w/ different -I manual keyed fName
# archiveProcessedFile() -- hoe doen ze dat bij Support
# --study warning and logging infrastructure
# emailPrognosis()
# extractRealized()
# emailRealized()
# =============================================================================

import argparse, calendar, csv, datetime, decimal, os, re, shutil, time
from openpyxl import cell, load_workbook, Workbook, styles
from openpyxl.cell import get_column_letter, coordinate_from_string
from openpyxl.cell import column_index_from_string

# on iMac24 at home
# importdir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts'
# archivedir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts/arch'
#
# on DellPC at work
importdir = 'H:\9_Issue_investigation\JIRA\SUM-10648-ING-LUX-HA-EPIC\SUM-10923\Baserun_infiles'
archivedir = 'H:\9_Issue_investigation\JIRA\SUM-10648-ING-LUX-HA-EPIC\SUM-10923\Baserun_outfiles_ARCH'
# missing is a default importdir and target output+archive directory/ now implicit geneation of outfiles in importdir...


def parseCmdLineInput():
    parser = argparse.ArgumentParser(description='\translate IR Exposure profile to IR Exposure replication trades')
    parser.add_argument("-I", dest="input_file", required=True)
    args = parser.parse_args()
    return args.input_file


def validateInputFile(fName):

    # valid file name---------------------------#
    fileYYYYMM = ''
    pattern = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(pattern, fName, re.I)
    
    try:
        if result is None:
            raise RuntimeError
    except RuntimeError:
        print ('Invalid fName:', fName, 'Expecting: replication*YYYYMM.xlsx')
        raise
    else:
        fileYYYYMM = (result.groupdict()['year']+result.groupdict()['month'])
        pass

    # file does not exist in importdir----------#
    absoluteImportFile = os.path.join(importdir, fName)
    try:
        if not os.path.isfile(absoluteImportFile):
            raise FileNotFoundError
    except FileNotFoundError as e:
        print ('File not found in import directory:', e)
        raise
    else:
        pass

    # file does not exist in archive dir(<>0Kb)-#
    os.chdir(archivedir)
    archList = os.listdir(archivedir)
    for file in archList:
        absoluteArchiveFile = os.path.join(archivedir, file)
        if os.pathfile.getsize(absoluteArchiveFile) > 0:
            continue
        archfileYYYYMM = ''
        try:
            match = re.search(pattern, file, re.I)
            if match is None:
                continue
            else:
                archfileYYYYMM = match.groupdict()['year']+match.groupdict()['month']
                if archfileYYYYMM == fileYYYYMM:
                    raise RuntimeError
        except RuntimeError:
            print('in archive dir exist files already processed for YYYYMM:',
                  fileYYYYMM, '>>> ', file,
                  time.strftime('%d/%m/%Y %H:%M:%S', time.gmtime(os.path.getmtime(file))))
            raise
        else:
            pass

    # read access establised to Sheet1---------#
    os.chdir(importdir)
    print('cwd', os.getcwd())
    wb = load_workbook(fName)
    sheetList = wb.get_sheet_names()
    try:
        ws1 = wb['Sheet1']
    except KeyError as e:
        print(e)
        print('No Sheet1 found', wb.get_sheet_names())
        exit()
    else:
        pass

    """
    # considered "too defensive" programming style - remove
    #
    # Sheet1:cellA1 matches up to fname YYYYMM-----#  deze vlieger gaat niet opgaan
    try:
        ws1 = wb['Sheet1']
        c = ws1.cell('A1')
        if c.value is None:
            raise AttributeError
        # to refine
        XLyyyymm = str((c.value.date().year))+str((c.value.date().month))
        if XLyyyymm == fileYYYYMM:
            pass
        else:
            raise RuntimeError
    except AttributeError:
        print('AttributeError: Sheet1:cellA1 is blank')
        raise()
        # to refine
    except RuntimeError:
        print('filecontent cell A1 does not match up with fileName YYYYMM')
        print(c, fileYYYYMM)
        raise()
    else:
        pass
    """

    # Sheet1 content limits equal range A1:CJ11-----#
    try:
        dim = ws1.calculate_dimension()
        if dim != 'A1:CJ11':
            raise RuntimeError
    except RuntimeError:
        print('Sheet1 content exceeding range A1:CJ11',
              'highest_column', cell.get_column_letter(ws1.get_highest_column()),
              'highest_row', ws1.get_highest_row())
        raise()

    # Sheet1 IR Exposure for zero% accounts with keys PR, RS, WS
    lubooks = (ws1.cell('B3').value, ws1.cell('B6').value, ws1.cell('B9').value)
    print('lubooks:', lubooks)
    try:
        for el in lubooks:
            if el not in ('PR', 'RS', 'WS'):

                raise NameError()
            else:
                continue
    except NameError:
        print('handled')
        raise()

    # other non-blank sheets exist > raise warning--#   misschien nok/ valt te zien
    wb = load_workbook(fName)
    wsList = wb.get_sheet_names()
    wsList.remove('Sheet1')
    for other in wsList:
        try:
            dim = wb[other].calculate_dimension()
            if dim != 'A1:A1':
                raise RuntimeError()
        except RuntimeError:
            print('Workbook contains other non empty sheets besides Sheet1!', other,
                  'highest_column', cell.get_column_letter(wb[other].get_highest_column()),
                  'highest_row', wb[other].get_highest_row())
            raise()


def processXls2Csv(myxlsFile):

    mappedFile = "PROGNOSIS_" + myxlsFile
    shutil.copyfile(myxlsFile, mappedFile)

    wbm = load_workbook(mappedFile)
    ws1 = wbm['Sheet1']

    ws_trades = wbm.create_sheet()
    ws_trades.title = 'Trades'

    csvFile = myxlsFile.replace(".xlsx", ".csv")
    csvFileObj = open(csvFile, 'w', newline="")
    csvData = csv.writer(csvFileObj)

    # -------------------write headerline------------------------------#
    headings = ["TradeDate",
                "MMType",
                "Ccy",
                "Start",
                "End",
                "Notional",
                "Index",
                "Rate",
                "Cpty",
                "Company",
                "Desk",
                "Book"]
    ws_trades.append(headings)
    csvData.writerow(headings)

    # -------------------set number format-----------------------------#
    _numberFormat = ws1['D3'].number_format

    # ---------------extract date from file name-----------------------#
    pattern = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(pattern, myxlsFile, re.I)
    YYYY = result.groupdict()['year']
    MM = result.groupdict()['month']

    # loop thru range with monthcolumn data to generate trades---------#
    # translate range boundaries into column_index equivalents
    processing_range = 'E3:CJ11'
    upperleft = (processing_range.partition(':')[0])
    lowerright = (processing_range.partition(':')[2])
    xy = coordinate_from_string(upperleft)
    startCol = column_index_from_string(xy[0])
    xy = coordinate_from_string(lowerright)
    maxCol = column_index_from_string(xy[0])
    # alternative---------------#
    # processing_range = 'E3:CJ11'
    # pattern = r'(?P<upperleftcolumn>^[A-Z]{1,2})(?P<upperleftrow>\d{1,2})(?P<delimiter>:)(?P<lowerrightcolumn>[A-Z]{1,2})(?P<lowerrightrow>\d{1,2}$)'
    # result = re.search(pattern, processing_range)
    # startcol = column_index_from_string(result.groupdict()['upperleftcolumn'])
    # maxcol   = column_index_from_string(result.groupdict()['lowerrightcolumn'])
    # alternative---------------#
    # startcol = 5 (icol for column E)
    # maxcol = ws1.get_highest_column()
    # --------------------------#

    for icol in range(startCol, maxCol+1):    # range(5,89) = column E to column CJ

        # -------------------populate trade attributes-------------#
        for irow in range(3, 10, 3):
            # default outfields
            Ccy = "EUR"
            Index = "FIXED"
            Cpty = "INTEINT"
            Company = "INGLU"
            Desk = "LUDESK"
            # other outfields population
            DD = calendar.monthrange(int(YYYY), int(MM))[1]
            date = YYYY + MM + "%d" %(DD)
            TradeDate = date
            Start = date

            cell = ws1.cell(row=irow, column=icol)
            if cell.value == 0:
                    continue
            elif cell.value < 0:
                    MMType = "DEPOSIT"
            else:
                    MMType = "LOAN"

            End = ws1.cell(row=2, column=icol).value
            Notional = "{0:.2f}".format(abs(ws1.cell(row=irow, column=icol).value))
            Rate = "{0:.4f}".format(ws1.cell(row=irow+1, column=icol).value/100)

            bookmap = {'PR': 'LUMMPB', 'RS': 'LUMMRE', 'WS': 'LUMMCB'}
            lubook = ws1.cell(row=irow, column=column_index_from_string('B')).value
            Book = bookmap[lubook]

            # populate trade dataline------------------------------#
            row = [TradeDate, MMType, Ccy, Start, End, Notional, Index, Rate, Cpty, Company, Desk, Book]
            ws_trades.append(row)
            csvData.writerow(row)

    del csvData
    csvFileObj.close()

    # -------------------generate InvCumGap profile--------------------#
    # add style color blue to calculated cell/range below
    ws1['A18'].value = 'maturing TOTAL'
    ws1['A19'].value = 'InvCumGap TOTAL'  # add bold style
    ws1['A21'].value = 'InvCumGap PR'
    ws1['A22'].value = 'InvCumGap RS'
    ws1['A23'].value = 'InvCumGap WS'
    ws1['A25'].value = 'InvCumGap TOTAL'
    ws1['A26'].value = 'Diff'

    for icol in range(startCol, maxCol+1):   # range(5,89) E=5 CJ=88
        # -----for each column enerate TOT total notional----------#
        cell = ws1.cell(row=18, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=3, column=icol).coordinate + "," + ws1.cell(row=6, column=icol).coordinate + "," + ws1.cell(row=9, column=icol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=19, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=18, column=icol).coordinate + ":" + ws1.cell(row=18, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=21, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=3, column=icol).coordinate + ":" + ws1.cell(row=3, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=22, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=6, column=icol).coordinate + ":" + ws1.cell(row=6, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=23, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=9, column=icol).coordinate + ":" + ws1.cell(row=9, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=25, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=21, column=icol).coordinate + "," + ws1.cell(row=22, column=icol).coordinate + "," + ws1.cell(row=23, column=icol).coordinate + ")"
        cell.number_format = _numberFormat

    cell = ws1.cell('E26')
    cell.value = "= " + ws1.cell('E19').value + "-" + ws1.cell('E25').value.strip("=") + " + 5 - 4 - 1"   # Diff=0 displays as blank cell
    print ("\'E26\'", cell.value)
    cell.number_format = _numberFormat

    wbm.save(mappedFile)


def iterateOverCoordinates():
    processing_range = 'E3:CJ11'
    print(processing_range)
    upperleft = (processing_range.partition(':')[0])
    lowerright = (processing_range.partition(':')[2])
    print(upperleft, lowerright)

    xy = coordinate_from_string(upperleft)
    print (xy)
    startcol = column_index_from_string(xy[0])
    print('startcol', startcol)

    xy = coordinate_from_string(lowerright)
    print (xy)
    endcol = column_index_from_string(xy[0])
    print('endcol', endcol)
    maxcol = endcol+1
    print('maxcol', maxcol)

    icol = 5
    print(icol)
    column = get_column_letter(icol)
    print(column)


def assignFormulaToCell():
    wb = load_workbook('Replication_201411.xlsx')
    ws1 = wb['Sheet1']
    _numberFormat = ws1['D3'].number_format

    cell_R17C5 = ws1.cell(row=17, column=5)
    print('ws1.cell(row=3, column=5).coordinate', ws1.cell(row=3, column=5).coordinate)
    print(ws1.cell(row=3, column=5).value)

    print('ws1.cell(row=6, column=5).coordinate', ws1.cell(row=6, column=5).coordinate)
    print(ws1.cell(row=6, column=5).value)

    print('ws1.cell(row9, column=5).coordinate', ws1.cell(row=9, column=5).coordinate)
    print(ws1.cell(row=9, column=5).value)

    cell_R17C5.value = "=SUM(" + ws1.cell(row=3, column=5).coordinate + "," + ws1.cell(row=6, column=5).coordinate + "," + ws1.cell(row=9, column=5).coordinate + ")"
    print ('cell_R17C5.value', type(cell_R17C5.value), cell_R17C5.value, cell_R17C5.internal_value)
    cell_R17C5.number_format = _numberFormat

    cell_R18C5 = ws1.cell(row=18, column=5)
    cell_R18C5.value = ws1.cell(row=3, column=5).value + ws1.cell(row=6, column=5).value + ws1.cell(row=9, column=5).value
    print('ws1.cell_R18C5(row=18, column=5).value', type(ws1.cell(row=18, column=5).value), ws1.cell(row=3, column=5).value)
    print (cell_R18C5.value)
    cell_R18C5.number_format = _numberFormat

    wb.save('Replication_201411_smalltest.xlsx')


def spielerei(fName):
    pass


def isValidFileName(fName):
    importdir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts'
    fileYYYYMM = ''
    pattern = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(pattern, fName, re.I)
    try:
        if result is None:
            raise RuntimeError
    except RuntimeError:
        print ('Invalid fName:', fName, 'Expecting: replication*YYYYMM.xlsx')
        raise
    else:
        fileYYYYMM = (result.groupdict()['year']+result.groupdict()['month'])
        return True

def exists(fName):
    # file exists in importdir--------------#
    importdir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts'
    absoluteImportFile = os.path.join(importdir, fName)
    try:
        if not os.path.isfile(absoluteImportFile):
            raise RuntimeError
    except RuntimeError as e:
        print ('File not found in import directory:', e)
        raise
    else:
        return True
    

def mainloop():
    print('mainloop started')
    inFile=parseCmdLineInput()
    validateInputFile(inFile)
    processXls2Csv(inFile)
    #iterateOverCoordinates()
    #assignFormulaToCell()
    #archiveProcessedFile()
    #--
    #emailPrognosis()
    #extractRealized()
    #emailRealized()
    print('mainloop done')
    pass


if __name__ == '__main__':
    """
    #assert os.path.isfile('PROGNOSIS_replication-201401.xlsx')
    """
    mainloop()

