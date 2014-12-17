

import os,csv,calendar,re,decimal,shutil
from openpyxl import load_workbook, Workbook, cell, styles
from optparse import OptionParser

parser = OptionParser()
parser.add_option("-I", dest="input_file", help="input file")
(options, args) = parser.parse_args()

if options.input_file == None:
	print("Input file is not defined. Please run '-h' to get more information.")
	exit()

inputFileName_original = re.findall('Replication.+\.xlsx', options.input_file)[0]
inputFileName_new = "PROGNOSIS_" + inputFileName_original
inputFile_new = options.input_file.replace(inputFileName_original,inputFileName_new)
shutil.copyfile(options.input_file, inputFile_new)
	
wb = load_workbook(inputFile_new)
try:
	ws_start = wb['Sheet1']
except KeyError as e:
	print(e)
	print(wb.get_sheet_names())
	exit()

for sheetName in wb.get_sheet_names():
	if sheetName == 'trade':
		ws_trade = wb.get_sheet_by_name(sheetName)
		wb.remove_sheet(ws_trade)
		break
	
ws_trade = wb.create_sheet()
ws_trade.title = 'trade'

csvFile = options.input_file.replace(".xlsx",".csv").replace(".xls","csv")
csvOpen = open(csvFile,'w',newline="")
csvData = csv.writer(csvOpen)

#---------------extract date from file name--------------
yearMonth = re.findall('Replication.+\.xlsx', options.input_file)[0][12:18]
year = yearMonth[0:4]
month = yearMonth[4:6]
day = calendar.monthrange(int(year),int(month))[1]
date = yearMonth + "%d" %(day)

#-------------------write the headers-------------#
headers = ["TradeDate","MMType","Ccy","Start","End","Notional","Index","Rate","Cpty","Company","Desk","Book"]
ws_trade.append(headers)
csvData.writerow(headers)
ws_start['B17'].value = 'TOT'
ws_start['A19'].value = 'InvCumGap TOT'
ws_start['A20'].value = 'InvCumGap PR'

#-------------------set number format----------------------------#
_numberFormat = ws_start['D3'].number_format

maxCol = ws_start.get_highest_column()
for icol in range(5,maxCol+1): #range(5,89) is from column E to column CJ
	#-------------------generate total notional---------------#
	cell = ws_start.cell(row=17,column=icol)
	cell.value = "=SUM(" + ws_start.cell(row=3,column=icol).coordinate + "," + ws_start.cell(row=6,column=icol).coordinate + "," + ws_start.cell(row=9,column=icol).coordinate + ")"
	cell.number_format = _numberFormat
	#-------------------generate trade scenarios---------------#
	for irow in range(3,10,3):
		cell = ws_start.cell(row=irow,column=icol)
		if cell.value == 0:
			continue
		elif cell.value < 0:
			MMType = "DEPOSIT"
		else:
			MMType = "LOAN"
		endDate = ws_start.cell(row=2,column=icol).value
		notional = "{0:.2f}".format(abs(ws_start.cell(row=irow,column=icol).value))
		rate = "{0:.4f}".format(ws_start.cell(row=irow+1,column=icol).value/100)
		book = ws_start.cell(row=irow,column=2).value
		if book == 'PR':
			lubook = 'LUMMPB'
		elif book == 'RS':
			lubook = 'LUMMRE'
		elif book == 'WS':
			lubook = 'LUMMCB'
		else:
			continue
		row = [date,MMType,"EUR",date,endDate,notional,"FIXED",rate,"INTEINT","INGLU","LUDESK",lubook]
		ws_trade.append(row)
		csvData.writerow(row)
del csvData
csvOpen.close()

#-------------------generate InvCumGap profile---------------#
for icol in range(5,maxCol+1): #range(5,89) is from column E to column CJ
	cell = ws_start.cell(row=19,column=icol)
	cell.value = "=SUM(" + ws_start.cell(row=17,column=icol).coordinate + ":" + ws_start.cell(row=17,column=maxCol).coordinate  + ")"
	cell.number_format = _numberFormat
	cell = ws_start.cell(row=20,column=icol)
	cell.value = "=SUM(" + ws_start.cell(row=3,column=icol).coordinate + ":" + ws_start.cell(row=3,column=maxCol).coordinate  + ")"
	cell.number_format = _numberFormat

wb.save(inputFile_new)
