# xlsx to csv.py
# v2 change to argparser
# v3 implement functions invoked under if __name__ == '__main__'
# v4 implement try: except: logic and replace row/col references by cell coordinates
# v5 introduce explicit importdir, outputdir, archivedir w/o implicitely assumed cwd
# v5 assumes CLI entry fore monthly changing -I filename ie no batch scheduling (not final)
# refactor for unix batch scheduling with "older" python xl libs xlrd, xlwt, xlutils
# delete blank Sheets other than Sheet1 in the ~_PROGNOSIS mappedfile
# check final xml outputs with 2 consecutive scripts run
# return args global for parseCmdLineInput \
# with defaults for optional arguments -D importdir and -A archivedir
# combine 2 scripts into one script
# implicit: output dir = script directory where also find xlsx input File
# implement from future import so that script could evt run in v2.4
# implement virtualenv/ explore
# implement Realized internal contracts ING LUX per Month dd, yyyy
# make script location agnostic with mapping file style almSwapUpload
# nose, pydev, coverage
# open questions:
# does Support run almSwapUpload each month w/ different -I manual keyed fName
# archiveProcessedFile() -- hoe doen ze dat bij Support
# --study warning and logging infrastructure
# emailPrognosis()
# extractRealized()
# emailRealized()
#archiveProcessedFile()
# =============================================================================

import argparse, calendar, csv, datetime, decimal, os, re, shutil, sys, time
from openpyxl import cell, load_workbook, styles, Workbook
from openpyxl.cell import coordinate_from_string, get_column_letter
from openpyxl.cell import column_index_from_string


def parseCmdLineInput():
    print('\n ..parseCmdLineInput')
    # on iMac24 at home
    # sourceDir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts'
    # targetDir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts/arch'
    
    # on DellPC at work
    sourceDir = 'H:\9_Issue_investigation\JIRA\SUM-10648-ING-LUX-HA-EPIC\SUM-10923\Baserun_infiles'
    targetDir = 'H:\9_Issue_investigation\JIRA\SUM-10648-ING-LUX-HA-EPIC\SUM-10923\Baserun_outfiles_ARCH'
    
    parser = argparse.ArgumentParser(description='\translate IR Exposure profile to deposit replication trades')
    parser.add_argument("-I", dest="infile", required=True,
                        help='*.xlsx input file')
    parser.add_argument("-SRC", dest="sourcedir", default=sourceDir,
                        help='XL feed directory')
    parser.add_argument("-DEST", dest="targetdir", default=targetDir,
                        help='outfile directory')
    args = parser.parse_args()
    return args


def validateInputFile(fName):
    print('\n ..validateInputFile')
    # valid file signature-----------------------#
    fileYYYYMM = ''
    patternXL = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(patternXL, fName, re.I)
    
    try:
        if result is None:
            raise RuntimeError
    except RuntimeError:
        print ('\nInvalid fName:', fName, 'Expecting: replication*YYYYMM.xlsx')
        raise
    else:
        fileYYYYMM = (result.groupdict()['year']+result.groupdict()['month'])
        pass

    # file should exist in importdir------------#
    absolutePathToInputFile = os.path.join(parserObj.sourcedir, fName)
    try:
        if not os.path.isfile(absolutePathToInputFile):
            raise FileNotFoundError
    except FileNotFoundError as e:
        print ('\nFile %s not found in import directory' % fName, e ,'\n')
        raise
    else:
        pass

    # duplicate processing check-----------------#
    # i.e. file (<>0Kb) with same signature does not already exist in target dir-#
    processedList = os.listdir(parserObj.targetdir)
    for file in processedList:
        # could evt only check for *.xlsx moved/archived files or *.xml outfiles same signature
        absoluteArchiveFile = os.path.join(parserObj.targetdir, file)

        # skip/ ignore 0KB files from unsuccesfull or incomplete runs
        if os.path.getsize(absoluteArchiveFile) == 0:
            continue
        archfileYYYYMM = ''
        patternCSV = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]csv$)'
        try:
            match = re.search(patternCSV, file, re.I)
            if match is None:
                continue
            else:
                archfileYYYYMM = match.groupdict()['year']+match.groupdict()['month']
                if archfileYYYYMM == fileYYYYMM:
                    raise RuntimeError
        except RuntimeError:
            print('\nduplicate processing! in archive dir exists file already processed for period YYYYMM:',
                  fileYYYYMM, '>>> ', '\n', file,
                  time.strftime('%d/%m/%Y %H:%M:%S', time.gmtime(os.path.getmtime(absoluteArchiveFile))))
            raise
        else:
            pass

    # read access establised to Sheet1---------#
    wb = load_workbook(absolutePathToInputFile)
    sheetList = wb.get_sheet_names()
    try:
        ws1 = wb['Sheet1']
    except KeyError as e:
        print('\nWorkbook only constains sheets', wb.get_sheet_names())
        print(e)
        exit()
    else:
        pass


    # Sheet1 content limits equal range A1:CJ11-----#
    try:
        dim = ws1.calculate_dimension()
        if dim != 'A1:CJ11':
            raise RuntimeError
    except RuntimeError:
        print('\nSheet1 content exceeding range A1:CJ11',
              'highest_column=', cell.get_column_letter(ws1.get_highest_column()),
              'highest_row=', ws1.get_highest_row())
        raise

    # Sheet1 IR Exposure for zero% accounts with keys PR, RS, WS
    lubooks = (ws1.cell('B3').value, ws1.cell('B6').value, ws1.cell('B9').value)
    try:
        for el in lubooks:
            if el not in ('PR', 'RS', 'WS'):
                raise NameError()
            else:
                continue
    except NameError:
        print('\n0% liabilities should either be: \
              \nPR private banking accounts \
              \nRS retail accounts \
              \nWS corporate accounts \
              \n!instead found>> ', el)
        raise

    # other non-blank sheets exist > raise warning--#   misschien nok/ valt te zien
    wb = load_workbook(absolutePathToInputFile)
    wsList = wb.get_sheet_names()
    wsList.remove('Sheet1')
    for other in wsList:
        try:
            dim = wb[other].calculate_dimension()
            if dim != 'A1:A1':
                raise RuntimeError()
        except RuntimeError:
            print('\nWorkbook contains other non empty sheets besides Sheet1!', '\n', 'e.g.', other,
                  'highest_column=', cell.get_column_letter(wb[other].get_highest_column()),
                  'highest_row=', wb[other].get_highest_row())
            raise


def processXls2Csv(fName):
    print('\n ..processXls2Csv')

    absolutePathToInputFile = os.path.join(parserObj.sourcedir, fName) 
    headTail =  os.path.splitext(fName)
    mappedfile = headTail[0] + '_PROGNOSIS' + headTail[1]
    absolutePathToMappedFile = os.path.join(parserObj.targetdir, mappedfile)
    shutil.copyfile(absolutePathToInputFile, absolutePathToMappedFile)
    
    wbm = load_workbook(absolutePathToMappedFile)
    ws1 = wbm['Sheet1']

    ws_trades = wbm.create_sheet()
    ws_trades.title = 'Trades'

    csvFile = fName.replace(".xlsx", ".csv")
    csvFileObj = open(os.path.join(parserObj.targetdir, csvFile), 'w', newline="")
    csvData = csv.writer(csvFileObj)

    # -------------------write headerline------------------------------#
    headings = ["TradeDate",
                "MMType",
                "Ccy",
                "Start",
                "End",
                "Notional",
                "Index",
                "Rate",
                "Cpty",
                "Company",
                "Desk",
                "Book"]
    ws_trades.append(headings)
    csvData.writerow(headings)

    # -------------------set number format-----------------------------#
    _numberFormat = ws1['D3'].number_format

    # ---------------extract date from file name-----------------------#
    patternXL = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(patternXL, fName, re.I)
    YYYY = result.groupdict()['year']
    MM = result.groupdict()['month']

    # loop thru range with monthcolumn data to generate trades---------#
    # translate range boundaries into column_index equivalents
    processing_range = 'E3:CJ11'
    upperleft = (processing_range.partition(':')[0])
    lowerright = (processing_range.partition(':')[2])
    xy = coordinate_from_string(upperleft)
    startCol = column_index_from_string(xy[0])
    xy = coordinate_from_string(lowerright)
    maxCol = column_index_from_string(xy[0])
    # alternative---------------#
    # processing_range = 'E3:CJ11'
    # pattern = r'(?P<upperleftcolumn>^[A-Z]{1,2})(?P<upperleftrow>\d{1,2})(?P<delimiter>:)(?P<lowerrightcolumn>[A-Z]{1,2})(?P<lowerrightrow>\d{1,2}$)'
    # result = re.search(pattern, processing_range)
    # startcol = column_index_from_string(result.groupdict()['upperleftcolumn'])
    # maxcol   = column_index_from_string(result.groupdict()['lowerrightcolumn'])
    # alternative---------------#
    # startcol = 5 (icol for column E)
    # maxcol = ws1.get_highest_column()
    # --------------------------#

    for icol in range(startCol, maxCol+1):    # range(5,89) = column E to column CJ

        # -------------------populate trade attributes-------------#
        for irow in range(3, 10, 3):
            # default outfields
            Ccy = "EUR"
            Index = "FIXED"
            Cpty = "INTEINT"
            Company = "INGLU"
            Desk = "LUDESK"
            # other outfields population
            DD = calendar.monthrange(int(YYYY), int(MM))[1]
            date = YYYY + MM + "%d" %(DD)
            TradeDate = date
            Start = date

            cell = ws1.cell(row=irow, column=icol)
            if cell.value == 0:
                    continue
            elif cell.value < 0:
                    MMType = "DEPOSIT"
            else:
                    MMType = "LOAN"

            End = ws1.cell(row=2, column=icol).value
            Notional = "{0:.2f}".format(abs(ws1.cell(row=irow, column=icol).value))
            Rate = "{0:.4f}".format(ws1.cell(row=irow+1, column=icol).value/100)

            bookmap = {'PR': 'LUMMPB', 'RS': 'LUMMRE', 'WS': 'LUMMCB'}
            lubook = ws1.cell(row=irow, column=column_index_from_string('B')).value
            Book = bookmap[lubook]

            # populate trade dataline------------------------------#
            row = [TradeDate, MMType, Ccy, Start, End, Notional, Index, Rate, Cpty, Company, Desk, Book]
            ws_trades.append(row)
            csvData.writerow(row)

    del csvData
    csvFileObj.close()

    # -------------------generate InvCumGap profile--------------------#
    # add style color blue to calculated cell/range below
    ws1['A18'].value = 'maturing TOTAL'
    ws1['A19'].value = 'InvCumGap TOTAL'  # add bold style
    ws1['A21'].value = 'InvCumGap PR'
    ws1['A22'].value = 'InvCumGap RS'
    ws1['A23'].value = 'InvCumGap WS'
    ws1['A25'].value = 'InvCumGap TOTAL'
    ws1['A26'].value = 'Diff'

    for icol in range(startCol, maxCol+1):   # range(5,89) E=5 CJ=88
        # -----for each column enerate TOT total notional----------#
        cell = ws1.cell(row=18, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=3, column=icol).coordinate + "," + ws1.cell(row=6, column=icol).coordinate + "," + ws1.cell(row=9, column=icol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=19, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=18, column=icol).coordinate + ":" + ws1.cell(row=18, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=21, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=3, column=icol).coordinate + ":" + ws1.cell(row=3, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=22, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=6, column=icol).coordinate + ":" + ws1.cell(row=6, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=23, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=9, column=icol).coordinate + ":" + ws1.cell(row=9, column=maxCol).coordinate + ")"
        cell.number_format = _numberFormat

        cell = ws1.cell(row=25, column=icol)
        cell.value = "=" + "SUM(" + ws1.cell(row=21, column=icol).coordinate + "," + ws1.cell(row=22, column=icol).coordinate + "," + ws1.cell(row=23, column=icol).coordinate + ")"
        cell.number_format = _numberFormat

    cell = ws1.cell('E26')
    cell.value = "= " + ws1.cell('E19').value + "-" + ws1.cell('E25').value.strip("=") + " + 5 - 4 - 1"   # Diff=0 displays as blank cell
    #print ("\'E26\'", cell.value)
    cell.number_format = _numberFormat

    wbm.save(absolutePathToMappedFile)


# embryonic functions for refactoring with regards to unit tests ###################
   
def hasValidFileSignature(fName):
    importdir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts'
    fileYYYYMM = ''
    patternXL = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(patternXL, fName, re.I)
    try:
        if result is None:
            raise RuntimeError
    except RuntimeError:
        print ('Invalid fName:', fName, 'Expecting: replication*YYYYMM.xlsx')
        raise
    else:
        fileYYYYMM = (result.groupdict()['year']+result.groupdict()['month'])
        print('file %s has valid File Signature %' % (fName, fileYYYYMM))
        return True


def exists(fName):
    # file exists in importdir--------------#
    importdir = '/Users/walter/Documents/GitExercises/hello-world/MijnPythonScripts'
    absoluteImportFile = os.path.join(importdir, fName)
    try:
        if not os.path.isfile(absoluteImportFile):
            raise RuntimeError
    except RuntimeError as e:
        print ('File not found in import directory:', e)
        raise
    else:
        return True
    
# </> end embryonic functions for refactoring with regards to unit tests ############

def spielerei():
    pass
    

def mainloop():
    now = time.strftime("%c")
    print('\n ..mainloop started', "current time %s"  % now)
    
    validateInputFile(parserObj.infile)
    processXls2Csv(parserObj.infile)
    
    now = time.strftime("%c")
    print('\n ..mainloop finished', "current time %s \n"  % now )


if __name__ == '__main__':
    parserObj = parseCmdLineInput()
    mainloop()

