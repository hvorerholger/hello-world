#! python3.4

sourceDirPath = 'H:\9_Issue_investigation\JIRA\SUM-10648-ING-LUX-HA-EPIC\SUM-10923\Baserun_infiles'
targetDirPath = 'H:\9_Issue_investigation\JIRA\SUM-10648-ING-LUX-HA-EPIC\SUM-10923\Baserun_outfiles_ARCH'
# check/assert parserObj.sourcedir and *.targetdir are filled with these default values if run options not explicited on command line!


def hasValidSignature(fName):
    print('\n ..hasValidateSignature')
    import re
    patternXL = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(patternXL, fName, re.I)
    try:
        if result is None:
            raise RuntimeError
    except RuntimeError:
        print ('Invalid fName:', fName, 'Expecting: replication*YYYYMM.xlsx')
        raise
    else:
        return True


def exists(fName, sourceDirPath):
    print('\n ..exists')
    import os
    absolutePathToInputFile = os.path.join(sourceDirPath, fName)
    try:
        if not os.path.isfile(absolutePathToInputFile):
            raise FileNotFoundError
    except FileNotFoundError as e:
        print ('\nFile %s not found in import directory %s ' % (fName, sourceDirPath), e ,'\n')
        raise
    else:
        return True


def noDuplicateProcessing(fName, targetDirPath):
    print('\n ..noDuplicateProcessing')
    import os, re, time

    patternXL = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]xlsx$)'
    result = re.search(patternXL, fName, re.I)
    fileYYYYMM = (result.groupdict()['year']+result.groupdict()['month'])
    
    processedList = os.listdir(targetDirPath)
    for file in processedList:
        # could evt only check for *.xlsx moved/archived files or *.xml outfiles same signature
        absoluteArchiveFile = os.path.join(targetDirPath, file)

        # skip/ ignore 0KB files from unsuccesfull or incomplete runs
        if os.path.getsize(absoluteArchiveFile) == 0:
            continue
        patternCSV = r'(?P<prefix>[. ]*replication)(?P<delimiter>[_| |-]?)(?P<year>20[1][4-9])(?P<month>0[1-9]|1[0-2])(?P<extension>[.]csv$)'
        try:
            match = re.search(patternCSV, file, re.I)
            if match is None:
                continue
            else:
                archfileYYYYMM = match.groupdict()['year']+match.groupdict()['month']
                if archfileYYYYMM == fileYYYYMM:
                    raise RuntimeError
        except RuntimeError:
            print('\n!duplicate processing! in archive dir exists file already processed for period YYYYMM=',
                  fileYYYYMM, ':', '\n', file,
                  time.strftime('%d/%m/%Y %H:%M:%S', time.gmtime(os.path.getmtime(absoluteArchiveFile))))
            raise
        else:
            pass
    return True


def canRead(fName, sheet):
    print('\n ..canRead')
    import os
    from openpyxl import load_workbook
    
    absolutePathToInputFile = os.path.join(sourceDirPath, fName)
    wb = load_workbook(absolutePathToInputFile)
    sheetList = wb.get_sheet_names()
    try:
        ws1 = wb[sheet]
    except KeyError as e:
        print('\nWorkbook "%s" only constains sheets "%s" ' % (fName, wb.get_sheet_names()))
        print(e)
        exit()
    else:
        return True


def hasValidDimension(fName, sheet):
    print('\n ..hasValidDimension')
    import os
    from openpyxl import cell, load_workbook
    
    absolutePathToInputFile = os.path.join(sourceDirPath, fName)
    wb = load_workbook(absolutePathToInputFile)
    ws1 = wb[sheet]
    try:
        dim = ws1.calculate_dimension()
        if dim != 'A1:CJ11':
            raise RuntimeError
    except RuntimeError:
        print('\nSheet1 content exceeding range A1:CJ11',
              'highest_column=', cell.get_column_letter(ws1.get_highest_column()),
              'highest_row=', ws1.get_highest_row())
        raise
    else:
        return True

def matchPortfolioKeys(fName, sheet, portfolioKeys):
    print('\n ..matchPortfolioKeys')
    import os
    from openpyxl import cell, load_workbook

    absolutePathToInputFile = os.path.join(sourceDirPath, fName)
    wb = load_workbook(absolutePathToInputFile)
    ws1 = wb[sheet]
    lubooks = (ws1.cell('B3').value, ws1.cell('B6').value, ws1.cell('B9').value)
    try:
        for el in lubooks:
            if el not in portfolioKeys:
                raise NameError()
            else:
                continue
    except NameError:
        print('\n0% liabilities should either be: \
              \nPR private banking accounts \
              \nRS retail accounts \
              \nWS corporate accounts \
              \n!instead found>> ', el)
        raise
    else:
        return True

def noOtherNonBlankSheetsApartFrom(fName, sheet):
    print('\n ..noOtherNonBlankSheetsApartFrom')
    import os
    from openpyxl import cell, load_workbook

    absolutePathToInputFile = os.path.join(sourceDirPath, fName)
    wb = load_workbook(absolutePathToInputFile)
    wsList = wb.get_sheet_names()
    wsList.remove(sheet)
    for other in wsList:
        try:
            dim = wb[other].calculate_dimension()
            if dim != 'A1:A1':
                raise RuntimeError()
        except RuntimeError:
            print('\nWorkbook %s contains other non empty sheets besides %s!' % (fName, sheet) , '\n', 'e.g.', other,
                  'highest_column=', cell.get_column_letter(wb[other].get_highest_column()),
                  'highest_row=', wb[other].get_highest_row())
            raise
    return True


if __name__ == "__main__":
    assert exists('Replication_201412.xlsx', sourceDirPath)
    assert hasValidSignature('Replication_201412.xlsx')
    assert noDuplicateProcessing('Replication_201411.xlsx', targetDirPath)
    assert canRead('Replication_201412.xlsx', 'Sheet1')
    assert hasValidDimension('Replication_201412.xlsx', 'Sheet1')
    assert matchPortfolioKeys('Replication_201412.xlsx','Sheet1',['PR','RS','WS'])
    assert noOtherNonBlankSheetsApartFrom('Replication_201412.xlsx', 'Sheet1')
    
    
    
